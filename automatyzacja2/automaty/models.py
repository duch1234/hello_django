from django.db import models
from django.utils import timezone
from django.core.validators import MinValueValidator, MaxValueValidator
from django.contrib import messages

import datetime
import re

from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
# Create your models here.


class Wygrzewanie(models.Model):
    STATUS_TESTU=(
        ('done','Wygrzany'),
        ("not done","Nie Wygrzany"),
    )
    
    STB_MODEL=(
        ('UHD','UHD'),
        ('WHD','WHD'),  
        ('Dakota', 'Dakota'),  
        ('WHD/UHD','WHD/UHD'),
    )
    
    title=models.CharField('Nazwa testu', max_length=300)
    titleClass=models.CharField('Nazwa klasy',max_length=300)
    author=models.CharField('Autor',max_length=50)
    type_stb=models.CharField('Typ dekodera',max_length=20,
                              choices=STB_MODEL,
                              default='WHD/UHD' )
    test_Status=models.CharField('Status',max_length=20,
                              choices=STATUS_TESTU,
                              default='done' )
    created=models.DateTimeField(auto_now=False, blank=True)
    def __str__(self):
        return self.title
    

    def importExcel(self):
        pass
        
    def generate_campaigne(self,request, queryset):
        
        response = HttpResponse(content_type='text/python')
        response['Content-Disposition'] = 'attachment; filename="Kampania.py"' 
        
        import_body=("import unittest \n"
                    "from xmlrunner import XMLTestRunner \n"
                    "from NewTvTesting.Utils import createAndGetXmlDirPath, writeTsSummaryToFiles\n"
                    "\n")
        
        test_suite_body=("\nif __name__ == '__main__':\n"
                        "\tsuite = unittest.TestSuite()\n")
        for i in queryset:
#             messages.add_message(request, messages.INFO, i)
            import_body=import_body+"from OPL_Testing."+str(i)+ " import "+ str(i) +"\n"
            test_suite_body=test_suite_body+"\tsuite.addTest("+str(i)+"(\"test\"))\n"
        
        test_suite_body=test_suite_body+("\n\trunner = XMLTestRunner(createAndGetXmlDirPath())\n"
                        "\tresult = runner.run(suite)\n"
                        "\twriteTsSummaryToFiles(result)\n"
                        "\tif not result.wasSuccessful():\n"
                        "\t\texit(1)\n")
        response.write(import_body+test_suite_body)
        return response
        
        
    generate_campaigne.short_description = u"Generuj kampanie"

class Konta_Testowe(models.Model):
    
    STATUS_KONTA=(
        ('active','Aktywne'),
        ("not active","Nie Aktywne"),
    )
    
    TECHNOLOGIA=(
        ('IPTV','IPTV'),
        ('DTH','DTH'),
    )
    accountNumber=models.PositiveIntegerField('Numer konta',
                                              validators=[MinValueValidator(1000000000,message="Za krótki numer konta. Sprawdz czy numer posiada 10 cyfr"),
                                                          MaxValueValidator(9999999999,message="za długi numer konta. Sprawdz czy numer posiada 10 cyfr")]
                                              )
    terminalNumbers=models.DecimalField('Ilość terminali',max_digits=1, decimal_places=0,
                                                validators=[MinValueValidator(1),
                                                MaxValueValidator(3),],
                                                default=3,
                                                )
    technology=models.CharField('Technologia', max_length=10,
                                choices=TECHNOLOGIA,
                                default='IPTV')
    
    accountStatus=models.CharField('Stan konta',max_length=20,
                                   choices=STATUS_KONTA,
                                   default="active")
    

def export_xlsx(modeladmin, request, queryset):

    #from openpyxl.cell import get_column_letter

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=mymodel.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Report"

    row_num = 0

    columns = [
        (u"ID", 15),
        (u"kartaName", 15),
        (u"serialNumber", 15),
    ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        #c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.pk,
            obj.kartaName,
            obj.serialNumber,
        ]
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
           # c.style.alignment.wrap_text = True

    wb.save(response)
    return response

export_xlsx.short_description = u"Export XLSX"
    

